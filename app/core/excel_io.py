import re
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook

REQUIRED_SHEETS = {"Input"}
CODE_PATTERN = re.compile(r"^([A-Za-z]+)")
GATEKEEPER_SUFFIXES = ("-G1", "-G2")


def load_workbook(path: str) -> Workbook:
    workbook = openpyxl.load_workbook(path, data_only=True)
    missing_sheets = REQUIRED_SHEETS - set(workbook.sheetnames)
    if missing_sheets:
        missing_names = ", ".join(sorted(missing_sheets))
        raise ValueError(
            f"Fehlende Sheets im Excel: {missing_names}. Bitte das bereitgestellte Template verwenden."
        )
    return workbook


def _normalize_code(value: Any) -> str | None:
    if not isinstance(value, str):
        return None

    normalized = value.strip()
    if not normalized:
        return None

    match = CODE_PATTERN.match(normalized)
    return match.group(1).upper() if match else None


def _parse_score(value: Any) -> float | None:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        normalized = value.strip().replace(",", ".")
        try:
            return float(normalized)
        except ValueError:
            return None

    return None


def _read_items(worksheet: Any) -> list[tuple[str, float]]:
    items: list[tuple[str, float]] = []

    for row_index in range(1, worksheet.max_row + 1):
        code = _normalize_code(worksheet.cell(row=row_index, column=3).value)
        score = _parse_score(worksheet.cell(row=row_index, column=5).value)
        if code and score is not None:
            items.append((code, score))

    return items


def _read_gatekeepers(worksheet: Any) -> dict[str, str]:
    gatekeepers: dict[str, str] = {}

    for row_index in range(1, worksheet.max_row + 1):
        gatekeeper_code = worksheet.cell(row=row_index, column=1).value
        gatekeeper_value = worksheet.cell(row=row_index, column=7).value

        if not isinstance(gatekeeper_code, str) or not isinstance(gatekeeper_value, str):
            continue

        normalized_code = gatekeeper_code.strip().upper()
        if normalized_code.endswith(GATEKEEPER_SUFFIXES):
            gatekeepers[normalized_code] = gatekeeper_value.strip()

    return gatekeepers


def read_inputs(workbook: Workbook) -> dict[str, Any]:
    worksheet = workbook["Input"]
    use_case_name = worksheet["B4"].value or "Unbenannter Use Case"

    items = _read_items(worksheet)
    if not items:
        raise ValueError("Keine Item-Bewertungen gefunden. Bitte im Tab 'Input' die Werte 1-5 ausfüllen.")

    return {
        "use_case_name": use_case_name,
        "items": items,
        "gatekeepers": _read_gatekeepers(worksheet),
    }
